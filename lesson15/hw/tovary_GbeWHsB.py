from openpyxl import load_workbook
# from openpyxl.drawing.image import Image

excel_file = load_workbook("result.xlsx")
excel_path = "result.xlsx"
page = excel_file["Лист1"]
headers = ['id', 'Название', 'Цена', 'Есть в наличии', 'Вес, кг', 'Картинка']
picture = ['cup.jpg', 'cup.jpg', 'cup.jpg', 'cup.jpg', 'cup.jpg']
products = {
    "A": {"id": [1, 2, 3, 4, 5]},
    "B": {"Название": ['Cup', 'Umbrella', 'Bike', 'Toy', 'Chair']},
    "C": {"Цена": [1000, 3000, 13000, 5000, 4000]},
    "D": {"Есть в наличии": ['yes', 'no', 'yes', 'no', 'yes']},
    "E": {"Вес, кг": [1, 3, 40, 0.5, 4]}
}

i = 0
for row in page["A1:F1"]:
    for cell in row:
        cell.value = headers[i]
        i += 1

b = []
for k in range(len(products)):
    a = []
    for key, qrt in products.items():
        for j in products[key]:
            a.append(products[key][j][k])
    b.append(a)

for row_idx in range(2, 7):  # ряды от 2 до 5 (включительно)
    for col_idx in range(1, 6):  # столбцы от 1 до 5 (включительно)
        page.cell(row=row_idx, column=col_idx, value=b[row_idx - 2][col_idx - 1])

# for idx in range(5):
#     img_name = picture[idx]
#     img = Image(img_name)
#     page.add_image(img, f"F{idx + 1}")
i = 0
for row in page["F2:F6"]:
    for cell in row:
        cell.value = picture[i]
        i += 1

excel_file.save(excel_path)
