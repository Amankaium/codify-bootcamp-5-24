# pip install openpyxl
from openpyxl import load_workbook


class ExcelReport:
    save_name = "result.xlsx"
    present_stat_name = "present_stat.xlsx"
    aoi = None  # "B2:D4"

    def __init__(self, path):
        self.excel_file = load_workbook(path)  # открываем файл
        self.page = self.excel_file[self.excel_file.sheetnames[0]]  # обращаемся к 1му листу в excel
        self.get_aoi()  # область интересов

    def get_aoi(self):
        letter = self.page["1"][-1].column_letter  # D
        number = self.page["A"][-1].row  # 4
        self.next_letter = self.get_next_letter(letter)  # E
        self.next_number = number + 1  # 5
        self.aoi = f"B2:{letter}{number}"
        return self.aoi

    def get_next_letter(self, letter):
        letters = "ABCDEFGHIGKLMNOPQRSTUVWXYZ"
        index = letters.index(letter)
        return letters[index + 1]

    def count_student_medium(self):
        self.page[f"{self.next_letter}1"] = "Средняя"
        for row in self.page[self.aoi]:
            sm = 0
            qty = 0
            for cell in row:
                sm += cell.value
                qty += 1
            row_value = cell.row
            self.page[f"{self.next_letter}{row_value}"] = sm / qty
        self.excel_file.save(self.save_name)

    def count_group_medium(self):
        sm = 0
        qty = 0
        for row in self.page[self.aoi]:
            for cell in row:
                sm += cell.value
                qty += 1

        self.page[f"{self.next_letter}{self.next_number}"] = sm / qty
        self.excel_file.save(self.save_name)

    def count_lesson_medium(self):
        self.page[f"A{self.next_number}"] = "Средняя"
        result = {}
        for row in self.page[self.aoi]:
            for cell in row:
                letter = cell.column_letter
                if letter not in result:
                    result[letter] = {"sum": 0, "qty": 0}

                    # {
                    # "B": {"sum": 275, "qty": 3},
                    # "C": {"sum": 220, "qty": 3},
                    # "D": {"sum": 230, "qty": 3}
                # }
                number = cell.value
                result[letter]["sum"] += number
                result[letter]["qty"] += 1

        for l in result:
            m = result[l]["sum"] / result[l]["qty"]
            self.page[f"{l}{self.next_number}"] = m
        self.excel_file.save(self.save_name)

    def count_present(self):
        self.page[f"{self.next_letter}1"] = "Посещаемость"
        for row in self.page[self.aoi]:
            present = 0
            for cell in row:
                if cell.value > 0:
                    present += 1
            row_value = cell.row
            self.page[f"{self.next_letter}{row_value}"] = present
        self.excel_file.save(self.present_stat_name)


file_1 = ExcelReport("../marks_2.xlsx")
# file_1.count_student_medium()
# file_1.count_group_medium()
# file_1.count_lesson_medium()
file_1.count_present()
