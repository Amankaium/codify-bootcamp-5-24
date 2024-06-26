import openpyxl
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
name_list = wb.sheetnames
print(name_list)


ws["A1"] = "ID"
ws["B1"] = "product_name"
ws["C1"] = "cost"
ws["D1"] = "weight"
ws["E1"] = "in stock"
ws["F1"] = "photo"


lst = [[1, "Cup", 200, 0.6, "Yes", "iq1000.png"],
       [2, "Umbrella", 300, 0.7, "No", "iq1000.png"],
       [3, "Toy", 400, 0.8, "Yes", "iq1000.png"],
       [4, "Table", 500, 0.9, "No", "iq1000.png"],
       [5, "Book", 600, 9, "No", "iq1000.png"]]


for row in lst:
    print(row)
    ws.append(row)

wb.save("product.xlsx")