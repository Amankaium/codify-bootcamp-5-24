from openpyxl import load_workbook
from openpyxl.styles import Font


class ExcelReport:
    present_stat_name = "present_stat.xlsx"
    aoi = None

    def __init__(self, path):
        self.excel_file = load_workbook(path)
        self.page = self.excel_file[self.excel_file.sheetnames[0]]
        self.get_aoi()

    def get_aoi(self):
        letter = self.page["1"][-1].column_letter
        number = self.page["A"][-1].row  # 4
        self.next_letter = self.get_next_letter(letter)
        self.next_number = number + 1
        self.aoi = f"B2:{letter}{number}"
        return self.aoi

    def get_next_letter(self, letter):
        letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        index = letters.index(letter)
        return letters[index + 1]

    def count_student_present(self):
        self.page[f"{self.next_letter}1"] = "Посещения"
        self.page[f"{self.next_letter}1"].font = Font(bold=True)
        for row in self.page[self.aoi]:
            qty = 0
            for cell in row:
                if cell.value != 0:
                    qty += 1
            row_value = cell.row
            self.page[f"{self.next_letter}{row_value}"] = qty
        self.excel_file.save(self.present_stat_name)

    def count_lesson_present(self):
        self.page[f"A{self.next_number}"] = "Посещения"
        self.page[f"A{self.next_number}"].font = Font(bold=True)
        result = {}
        for row in self.page[self.aoi]:
            for cell in row:
                letter = cell.column_letter
                if letter not in result:
                    result[letter] = {"qty": 0}
                if cell.value != 0:
                    result[letter]["qty"] += 1

        for l in result:
            presents_qty = result[l]["qty"]
            self.page[f"{l}{self.next_number}"] = presents_qty
        self.excel_file.save(self.present_stat_name)


file_1 = ExcelReport("../marks_2.xlsx")

file_1.count_student_present()
file_1.count_lesson_present()
