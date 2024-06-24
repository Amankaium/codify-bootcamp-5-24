# pip install openpyxl
from openpyxl import load_workbook


class ExcelReport:
    save_name = "result.xlsx"
    
    def __init__(self, path):
        self.excel_file = load_workbook(path)  # открываем файл
        self.page = self.excel_file[self.excel_file.sheetnames[0]]  # обращаемся к 1му листу в excel
        
    def count_student_medium(self):
        self.page["E1"] = "Средняя"
        for row in self.page["B2:D4"]:
            sm = 0
            qty = 0
            for cell in row:
                sm += cell.value
                qty += 1
            row_value = cell.row
            self.page[f"E{row_value}"] = sm / qty
        self.excel_file.save(self.save_name)

file_1 = ExcelReport("marks_2.xlsx")
file_1.count_student_medium()
